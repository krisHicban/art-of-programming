# main_script_simple.py

import os
import random
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# --- Configuration ---
EXCEL_FILE_PATH = 'angajati_data_simple.xlsx'
NUM_EMPLOYEES = 1000


# --- Helper function to create the Excel file with random data ---
def create_and_populate_excel_if_not_exists(filepath, num_records):
    """
    Checks if the Excel file exists. If not, it creates it and populates it
    with random employee data using the openpyxl library.
    """
    if os.path.exists(filepath):
        print(f"File '{filepath}' already exists. Reading data from it.")
        return

    print(f"File '{filepath}' not found. Generating new data for {num_records} employees...")

    # Create a new workbook and select the active sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Angajati"

    # Define headers
    headers = ['Angajat', 'Salary', 'Experience']
    sheet.append(headers)

    # Sample data for randomization
    first_names = ["Andrei", "Mihai", "Ion", "Vasile", "Stefan", "Elena", "Maria", "Ioana", "Ana", "Cristina"]
    last_names = ["Popescu", "Ionescu", "Popa", "Radu", "Dumitru", "Stoica", "Georgescu", "Matei", "Constantin"]

    # Generate and append random data
    for _ in range(num_records):
        employee_name = f"{random.choice(first_names)} {random.choice(last_names)}"
        salary = random.randint(3000, 15000)
        experience = random.randint(0, 15)
        sheet.append([employee_name, salary, experience])

    # Save the file
    try:
        workbook.save(filepath)
        print(f"Successfully created and populated '{filepath}'.")
    except Exception as e:
        print(f"An error occurred while creating the Excel file: {e}")
        exit()


# --- Core function for bonus calculation ---
def calculeaza_bonus(salariu, experienta):
    """
    Calculează bonusul pe baza experienței.
    Returnează: valoarea bonusului
    """
    # Ensure inputs are valid numbers before calculation
    if not isinstance(salariu, (int, float)) or not isinstance(experienta, (int, float)):
        return 0

    if experienta >= 5:
        return salariu * 0.20  # 20% bonus
    elif experienta >= 3:
        return salariu * 0.15  # 15% bonus
    elif experienta >= 1:
        return salariu * 0.10  # 10% bonus
    else:
        return 0  # Fără bonus


# --- Main script execution ---
if __name__ == "__main__":
    # Step 1: Ensure the Excel file exists. Create and populate it if it doesn't.
    create_and_populate_excel_if_not_exists(EXCEL_FILE_PATH, NUM_EMPLOYEES)

    # Step 2: Read from Excel and calculate the bonus
    try:
        print("\nLoading Excel workbook to process...")
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        sheet = workbook.active
    except FileNotFoundError:
        print(f"Error: The file '{EXCEL_FILE_PATH}' could not be found.")
        exit()
    except Exception as e:
        print(f"An error occurred while loading the Excel file: {e}")
        exit()

    print("Workbook loaded. Calculating bonuses and updating the sheet...")

    # Add the 'Bonus' header in the 4th column (Column D)
    header_cell = sheet.cell(row=1, column=4)
    header_cell.value = 'Bonus'

    # Iterate over all rows in the sheet, starting from the second row to skip the header
    # `sheet.iter_rows` is efficient as it doesn't load all rows into memory at once
    for row_index, row_cells in enumerate(sheet.iter_rows(min_row=2), start=2):
        # The `row_cells` is a tuple of cells for the current row
        # Column A (index 0) is Angajat, B (1) is Salary, C (2) is Experience
        nume = row_cells[0].value
        salariu = row_cells[1].value
        experienta = row_cells[2].value

        # Calculate the bonus
        bonus = calculeaza_bonus(salariu, experienta)

        # Write the calculated bonus to the 4th column of the current row
        bonus_cell = sheet.cell(row=row_index, column=4)
        bonus_cell.value = round(bonus, 2)
        # Optional: Apply a number format for currency
        bonus_cell.number_format = '#,##0.00 "RON"'

        # Print a few examples to show the script is working
        if row_index <= 6:  # Print for the first 5 employees (row 2 to 6)
            print(f"Salariul lui {nume} este {salariu} RON, "
                  f"iar bonusul calculat este {bonus_cell.value} RON.")

    # Step 3: At the end, save the modified workbook back to the same file
    try:
        print(f"\nSaving updated data back to '{EXCEL_FILE_PATH}'...")
        workbook.save(EXCEL_FILE_PATH)
        print("Successfully updated the Excel file with the 'Bonus' column.")
    except Exception as e:
        print(f"An error occurred while saving the updated Excel file: {e}")